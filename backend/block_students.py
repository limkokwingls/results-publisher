import os
from dataclasses import asdict
from turtle import st

import openpyxl
from blocked import get_blocked_students
from firebase_admin import credentials, firestore
from models import CourseGrade, Student
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich import print
from rich.console import Console
from utils import is_number, to_float

console = Console()
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def get_data_files(dir: str):
    files = []
    dir = os.path.join(BASE_DIR, dir)
    for file in os.listdir(dir):
        if file.endswith(".xlsx"):
            files.append(os.path.join(dir, file))
    return files


def std_number_column(sheet: Worksheet):
    for col in sheet.iter_cols():
        for c in col:
            cell: Cell = c
            if cell.value == "StudentID":
                return cell.column
    return 3


def get_last_data_column(sheet: Worksheet):
    for col in sheet.iter_cols():
        for c in col:
            cell: Cell = c
            if cell and cell.value and "remark" in str(cell.value).lower():
                return cell.column
    return sheet.max_column  # Return the last column if "remark" is not found


def highlight_blocked(sheet: Worksheet, std_numbers: list[str]):
    std_number_col = std_number_column(sheet)
    last_data_col = get_last_data_column(sheet)
    for row in sheet.iter_rows(
        min_row=2, max_col=last_data_col
    ):  # Start from row 2 to skip header
        cell = row[std_number_col - 1]  # Adjust for 0-based index
        if str(cell.value) in std_numbers:
            fill_row(row, last_data_col)


def fill_row(row, last_data_col: int):
    black_fill = openpyxl.styles.PatternFill(
        start_color="000000", end_color="000000", fill_type="solid"
    )
    for cell in row[:last_data_col]:
        cell.fill = black_fill


def main():
    files = get_data_files("data")
    students = get_blocked_students()

    print("Blocked Students: ", len(students))

    for i, file in enumerate(files):
        file_name = os.path.basename(file)
        print(f"{i+1}/{len(files)}) {file_name}")
        workbook: Workbook = openpyxl.load_workbook(file)
        for i, ws in enumerate(workbook):
            print(f"Sheet {i+1}: {ws.title}")
            highlight_blocked(ws, students)

        # Save the modified workbook
        workbook.save(file)


if __name__ == "__main__":
    main()
