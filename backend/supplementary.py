import os
import re
from types import NoneType

import openpyxl
from base import Base, Session, engine
from models import CourseGrade, Faculty, Program, Student, StudentClass
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich import print
from rich.console import Console
from sqlalchemy.orm import SessionTransaction
from utils import is_number, to_float, to_int

console = Console()
Base.metadata.create_all(engine)
session = Session()
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def get_std_column(sheet: Worksheet):
    for col in sheet.iter_cols():
        for c in col:
            cell: Cell = c
            if cell.value == "StudentID":
                return cell.column
    return 3


def get_remarks_column(sheet: Worksheet):
    for col in sheet.iter_cols():
        for c in col:
            cell: Cell = c
            if "faculty remark" in str(cell.value).lower():
                return cell.column
    raise Exception("Remarks column not found")


def get_std_name_column(sheet: Worksheet):
    for col in sheet.iter_cols():
        for c in col:
            cell: Cell = c
            if cell.value == "Name":
                return cell.column
    return 2


def read_worksheet(sheet: Worksheet):
    course_name = None
    for row in sheet.iter_rows():
        for cell in row:
            if type(cell.value) is str and "module" in cell.value.lower():
                course_name = cell.value
                # remove "Module" or "MODULE" or "module" from course name
                course_name = re.sub(r"module", "", course_name, flags=re.IGNORECASE)
                course_name = re.sub(r":", "", course_name)
                course_name = course_name.strip()
                break
        # get first cell in row
        noValue = row[0].value
        if is_number(noValue) and row[2].value:
            std_no = str(row[2].value)
            # remove white spaces
            std_no = re.sub(r"\s+", "", std_no)
            student = session.query(Student).filter_by(no=std_no).first()
            if not student:
                student = Student(
                    no=std_no,
                    name=row[1].value,
                    remarks=row[6].value,
                    student_class_id=-1,
                )
                session.add(student)
            grades = CourseGrade(
                name=course_name,
                code="",
                grade=row[6].value,
                points="0.0",
                marks=row[5].value,
                student_no=student.no,
            )
            session.add(grades)
            print(student, grades)

            session.commit()


def get_data_files(dir: str):
    files = []
    dir = os.path.join(BASE_DIR, dir)
    for file in os.listdir(dir):
        if file.endswith(".xlsx"):
            files.append(os.path.join(dir, file))
    return files


def delete_everything():
    session.query(CourseGrade).delete()
    session.query(Student).delete()
    session.query(StudentClass).delete()
    session.query(Program).delete()
    session.query(Faculty).delete()
    session.commit()


def main():
    with console.status("Clearing database..."):
        delete_everything()
    files = get_data_files("data")
    for i, file in enumerate(files):
        file_name = file.split("\\")[-1]
        print(f"{i+1}/{len(files)}) {file_name}")
        workbook: Workbook = openpyxl.load_workbook(file)
        for i, ws in enumerate(workbook):
            sheet: Worksheet = ws
            with console.status(
                f"{i + 1}/{len(workbook.worksheets)} Processing '{sheet.title}'..."
            ):
                print(sheet.title)
                read_worksheet(sheet)
    print("Done!")


if __name__ == "__main__":
    main()
