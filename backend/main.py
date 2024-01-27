import os
from dataclasses import asdict

import firebase_admin
import openpyxl
from firebase_admin import credentials, firestore
from models import CourseGrade, Student
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich import print
from rich.console import Console
from utils import is_number, to_float

console = Console()
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def get_std_column(sheet: Worksheet):
    for col in sheet.iter_cols():
        for c in col:
            cell: Cell = c
            if cell.value == "StudentID":
                return cell.column
    return 3


def get_remarks_column(sheet: Worksheet):
    for col in sheet.iter_cols():
        for c in col:
            cell: Cell = c
            if "faculty remark" in str(cell.value).lower():
                return cell.column
    raise Exception("Remarks column not found")


def get_std_name_column(sheet: Worksheet):
    for col in sheet.iter_cols():
        for c in col:
            cell: Cell = c
            if cell.value == "Name":
                return cell.column
    return 2


def get_course_rows(sheet: Worksheet):
    courses = {}
    for col in sheet.iter_cols():
        for c in col:
            cell: Cell = c
            if cell.value == "Mk":
                col_i = cell.col_idx
                mk_label_i = cell.row  # The row index for "MK"
                code_cell: Cell = sheet.cell(mk_label_i - 2, col_i)
                name = None
                i = 3
                try:
                    while i < 10:
                        name_cell = sheet.cell(mk_label_i - i, col_i)
                        name = name_cell.value
                        if name:
                            break
                        i += 1
                except Exception as e:
                    print(e)

                # courses[grade_cell.value] = grade_cell.column
                courses[code_cell.column] = {
                    "code": code_cell.value,
                    "name": name,
                }
    return courses


def get_students_with_rows(sheet: Worksheet) -> dict[int, Student]:
    data = {}
    std_col = get_std_column(sheet)
    std_name_col = get_std_name_column(sheet)

    for i, row in enumerate(sheet.iter_rows()):
        num = None
        name = None
        for _cell in row:
            cell: Cell = _cell
            if cell.column == std_col:
                if is_number(cell.value):
                    num = cell.value
            if cell.column == std_name_col:
                name = cell.value
        if num and name:
            data[i + 1] = Student(name=name, no=num)

    return data


cred = credentials.Certificate("serviceAccountKey.json")
app = firebase_admin.initialize_app(cred)
db = firestore.client(app)


def save_to_firestore(students: dict[int, Student]):
    print("Saving to database")
    size = len(students)
    for i, std in enumerate(students.values()):
        print(f"{i+1}/{size} Saving {std.name} ({std.no})...")
        std_ref = db.collection("students").document(str(int(std.no)))
        std_ref.set(asdict(std))


def get_students_with_grades(sheet: Worksheet):
    marks_dict = get_course_rows(sheet)
    students = get_students_with_rows(sheet)
    remarks_col = get_remarks_column(sheet)

    student_dict = {}

    for student_col, std in students.items():
        for mark_col, course in marks_dict.items():
            try:
                marks_cell: Cell = sheet.cell(student_col, mark_col)
                grade_cell: Cell = sheet.cell(student_col, mark_col + 1)
                points_cell: Cell = sheet.cell(student_col, mark_col + 2)
                if is_number(marks_cell.value):
                    marks = to_float(marks_cell.value)
                    grade = str(grade_cell.value)
                    points = to_float(points_cell.value)
                    student = student_dict.get(std.no)
                    if not student:
                        student = Student(name=std.name, no=int(std.no))
                        student_dict[std.no] = student

                    course_grade = CourseGrade(
                        code=course["code"],
                        name=course["name"],
                        marks=marks,
                        grade=grade,
                        points=points,
                    )
                    student.course_grades.append(course_grade)

                    remarks_cell: Cell = sheet.cell(student_col, remarks_col)
                    if remarks_cell.value:
                        remarks = str(remarks_cell.value)
                        student.remarks = remarks
            except Exception as e:
                print(e)
    return student_dict


def get_data_files(dir: str):
    files = []
    dir = os.path.join(BASE_DIR, dir)
    for file in os.listdir(dir):
        if file.endswith(".xlsx"):
            files.append(os.path.join(dir, file))
    return files


def main():
    files = get_data_files("data")
    students: dict[int, Student] = {}
    for i, file in enumerate(files):
        file_name = file.split("\\")[-1]
        print(f"{i+1}/{len(files)}) {file_name}")
        workbook: Workbook = openpyxl.load_workbook(file)
        for i, ws in enumerate(workbook):
            sheet: Worksheet = ws
            with console.status(
                f"{i + 1}/{len(workbook.worksheets)} Processing '{sheet.title}'..."
            ):
                print(sheet.title)
                if sheet.title and "sheet" in sheet.title.lower():
                    print(f"Skipping {sheet.title}...")
                    continue
                # student_class = create_student_class(sheet)
                students.update(get_students_with_grades(sheet))

    save_to_firestore(students)
    print("Done!")


if __name__ == "__main__":
    main()
