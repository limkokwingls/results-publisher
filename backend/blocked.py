import os

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rich import print
from rich.console import Console
from utils import is_number
from db import get_db, cleanup_db

console = Console()
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def get_std_no_column(sheet: Worksheet):
    for col in sheet.iter_cols():
        for c in col:
            cell: Cell = c
            if "Student ID" in str(cell.value):
                return cell.column - 1
    return 1


def get_student_numbers(sheet: Worksheet) -> list[str]:
    student_numbers = []
    col_number = get_std_no_column(sheet)
    print("Reading Sheet ", sheet.title, "Column Number: ", col_number)
    for row in sheet.iter_rows():
        cell: Cell = row[col_number]
        if is_number(cell.value):
            student_numbers.append(as_str(cell.value))

    return student_numbers


def as_str(value) -> str:
    val = str(value).strip()
    if val.endswith(".0"):
        return val[:-2]
    return val


def get_blocked_students():
    students = []
    workbook: Workbook = openpyxl.load_workbook("Blocked.xlsx")
    for sheet in workbook:
        student_numbers = get_student_numbers(sheet)
        students.extend(student_numbers)
    return students


def unblock(student_number: str) -> None:
    app, db = get_db()
    db.collection("students").document(str(student_number)).set(
        {"is_blocked": False}, merge=True
    )
    cleanup_db(app)


def main() -> None:
    app, db = get_db()
    student_numbers = get_blocked_students()
    for i, num in enumerate(student_numbers):
        db.collection("students").document(str(num)).set(
            {"is_blocked": True}, merge=True
        )
        print(f"{i + 1}/{len(student_numbers)}) {num} blocked")
    print("Done!")
    cleanup_db(app)


if __name__ == "__main__":
    main()
